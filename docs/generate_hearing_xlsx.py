# -*- coding: utf-8 -*-
"""
株式会社ヴィレッジ HP ヒアリングシート生成スクリプト
各シート = 各ページ／確認項目。現状のドラフト内容 + ヒアリング記入欄。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
wb.remove(wb.active)

# --- スタイル定義（ブランドカラー：青×緑） ---
THIN = Side(border_style="thin", color="C7CDD4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT = Font(name="游ゴシック", size=18, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="0369A1")   # sky-700

SECTION_FONT = Font(name="游ゴシック", size=12, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="334155")  # slate-700

LABEL_FONT = Font(name="游ゴシック", size=11, bold=True, color="1E293B")
LABEL_FILL = PatternFill("solid", fgColor="E2E8F0")    # slate-200

CURRENT_FONT = Font(name="游ゴシック", size=10, color="475569")
CURRENT_FILL = PatternFill("solid", fgColor="F1F5F9")  # slate-100

INPUT_FONT = Font(name="游ゴシック", size=11, color="0F172A")
INPUT_FILL = PatternFill("solid", fgColor="FFFFFF")

NOTE_FONT = Font(name="游ゴシック", size=9, color="64748B", italic=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def setup_sheet(ws, page_name, page_path):
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 52
    ws.column_dimensions["E"].width = 4

    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = f"【{page_name}】ヒアリングシート   ( {page_path} )"
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = CENTER
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:D3")
    c = ws["B3"]
    c.value = "中央列＝現在の制作内容（ドラフト） ／ 右列＝ヒアリング記入欄（修正・ご希望をこちらにご記入ください）"
    c.font = NOTE_FONT
    c.alignment = CENTER
    ws.row_dimensions[3].height = 22

    for cell, val in [("B4", "項目"), ("C4", "現在の制作内容（ドラフト）"), ("D4", "ヒアリング記入欄")]:
        ws[cell] = val
        ws[cell].font = SECTION_FONT
        ws[cell].fill = SECTION_FILL
        ws[cell].alignment = CENTER
        ws[cell].border = BORDER
    ws.row_dimensions[4].height = 26
    return 5


def add_section(ws, row, title):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2, value=f"■ {title}")
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24
    return row + 1


def add_row(ws, row, label, current, height=None):
    lc = ws.cell(row=row, column=2, value=label)
    lc.font = LABEL_FONT; lc.fill = LABEL_FILL; lc.alignment = LEFT_WRAP; lc.border = BORDER
    cc = ws.cell(row=row, column=3, value=current)
    cc.font = CURRENT_FONT; cc.fill = CURRENT_FILL; cc.alignment = LEFT_WRAP; cc.border = BORDER
    ic = ws.cell(row=row, column=4, value="")
    ic.font = INPUT_FONT; ic.fill = INPUT_FILL; ic.alignment = LEFT_WRAP; ic.border = BORDER
    if height:
        ws.row_dimensions[row].height = height
    else:
        lines = max(1, len(str(current)) // 30) + str(current).count("\n")
        ws.row_dimensions[row].height = max(22, min(180, 18 + lines * 14))
    return row + 1


def add_spacer(ws, row):
    ws.row_dimensions[row].height = 6
    return row + 1


# ===========================================
# 00_表紙（あいさつ文・記入方法）
# ===========================================
ws = wb.create_sheet("00_表紙")
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 74

ws.merge_cells("B2:C2")
c = ws["B2"]
c.value = "株式会社ヴィレッジ 様　ホームページ ヒアリングシート"
c.font = Font(name="游ゴシック", size=20, bold=True, color="FFFFFF")
c.fill = TITLE_FILL
c.alignment = CENTER
ws.row_dimensions[2].height = 48

# あいさつ文
ws.merge_cells("B4:C4")
c = ws["B4"]
c.value = "ごあいさつ"
c.font = SECTION_FONT; c.fill = SECTION_FILL
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[4].height = 24

greeting = (
    "株式会社ヴィレッジ 中川 基次 様\n\n"
    "この度はホームページ制作のご依頼をいただき、誠にありがとうございます。\n"
    "お預かりしたお名刺の情報をもとに、まずはサイトのたたき台（ドラフト）を作成いたしました。\n\n"
    "本シートは、その内容を一緒に確認・調整するための資料です。\n"
    "各ページの「現在の制作内容」をご覧いただき、直したい点・追加したい点を"
    "右側の記入欄にご記入ください。難しく考えず、気づいたことをメモ感覚で書いていただければ大丈夫です。\n\n"
    "ご記入後、本ファイルをそのままメールにてご返信ください。内容を反映して仕上げてまいります。\n"
    "何卒よろしくお願い申し上げます。"
)
ws.merge_cells("B5:C5")
c = ws["B5"]
c.value = greeting
c.font = Font(name="游ゴシック", size=11, color="0F172A")
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws.row_dimensions[5].height = 210

instructions = [
    ("ご記入方法", [
        "① 画面下部のシートタブ（TOP / 事業内容 / 会社概要 …）を切り替えて各ページをご確認ください。",
        "② 各シートは「現在の制作内容（ドラフト）」と「ヒアリング記入欄」の2列構成です。",
        "③ 右側の白いセルに、修正・ご希望・正しい情報をご記入ください。",
        "④ そのままでOKな項目は空欄のままで結構です。",
        "⑤ ご記入後、本ファイルをメールにてご返信ください。",
    ]),
    ("確認用URL（現在のドラフト）", [
        "制作中のサイトはこちらでご確認いただけます（公開準備が整い次第URLをお送りします）：",
        "　▶ 確認用URL：______________________________（担当より別途ご連絡）",
        "　※ スマートフォン・パソコン、どちらからでもご覧いただけます。",
    ]),
    ("シート構成", [
        "00_表紙        … 本シート（ごあいさつ・記入方法）",
        "01_TOP         … トップページ（一番最初に表示される画面）",
        "02_事業内容     … 6つの事業の説明",
        "03_選ばれる理由  … 会社の強み",
        "04_会社概要ほか  … 会社情報・ご提供写真・デザインのご希望をまとめて",
        "05_お問い合わせ  … 電話・メールの表示内容",
    ]),
    ("ご提出先", [
        "ホームページ制作担当",
        "Email：______________________（担当メールアドレス）",
    ]),
]

row = 7
for header, lines in instructions:
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c = ws.cell(row=row, column=2, value=header)
    c.font = Font(name="游ゴシック", size=12, bold=True, color="FFFFFF")
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 1
    for line in lines:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        c = ws.cell(row=row, column=2, value=line)
        c.font = Font(name="游ゴシック", size=11, color="0F172A")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)
        ws.row_dimensions[row].height = 20
        row += 1
    row += 1

# ===========================================
# 01_TOP
# ===========================================
ws = wb.create_sheet("01_TOP")
row = setup_sheet(ws, "TOPページ", "トップ画面")

row = add_section(ws, row, "メインビジュアル（一番上の大きな画面）")
row = add_row(ws, row, "キャッチコピー", "解体から、まちの未来をつくる。")
row = add_row(ws, row, "紹介文",
    "北九州市門司区を拠点に、建物解体工事・土地開発・不動産・リフォームまでをワンストップでご提供。"
    "安全・丁寧・スピーディーな施工で、地域の暮らしと未来を支えます。")
row = add_row(ws, row, "背景写真", "（現状：仮のデザイン背景）\n※ 実際の現場・重機などの写真に差し替え可能です。ご希望をご記入ください。")
row = add_row(ws, row, "上部の表示ラベル", "北九州市門司区 / 解体・不動産・リフォーム")
row = add_row(ws, row, "3つの数字表示", "対応エリア：北九州　／　見積り：無料　／　有資格者：在籍\n※ 変更・追加したい数字（施工実績◯件 等）があればご記入ください。")

row = add_spacer(ws, row)
row = add_section(ws, row, "ボタン（お客様が押す部分）")
row = add_row(ws, row, "メインボタン", "無料で相談・お見積り")
row = add_row(ws, row, "サブボタン", "事業内容を見る")

# ===========================================
# 02_事業内容
# ===========================================
ws = wb.create_sheet("02_事業内容")
row = setup_sheet(ws, "事業内容", "6つの事業")

services = [
    ("建物解体工事", "戸建住宅から店舗・ビルまで、木造・鉄骨・RC造に対応。近隣に配慮した安全管理と適正な分別解体を徹底します。"),
    ("解体業コンサルタント", "解体計画の立案から届出・許認可、見積の精査まで。一級土木施工管理技士が最適なプランをご提案します。"),
    ("土地開発", "造成・整地・インフラ整備まで、土地の価値を最大化する開発をトータルでサポートします。"),
    ("不動産売買", "解体後の土地活用や売却、物件の買取までワンストップ。地域に精通したスタッフが親身に対応します。"),
    ("リフォーム事業", "水回りから外装・内装まで、住まいの困りごとを解決。暮らしやすさと資産価値の向上をお手伝いします。"),
    ("リノベーション事業", "中古住宅を新たな価値へ。ライフスタイルに合わせた空間づくりをデザインから施工まで一貫して行います。"),
]
for i, (title, desc) in enumerate(services, start=1):
    row = add_section(ws, row, f"事業{i}　{title}")
    row = add_row(ws, row, "説明文", desc)
    row = add_row(ws, row, "修正・補足", "", height=36)
    row = add_spacer(ws, row)

row = add_section(ws, row, "その他")
row = add_row(ws, row, "追加したい事業", "", height=40)
row = add_row(ws, row, "対応可能エリア", "※ 対応できる市区町村があればご記入ください（例：北九州市全域・下関市 等）", height=36)
row = add_row(ws, row, "保有資格・許可", "一級土木施工管理技士\n※ 建設業許可番号・解体工事業登録番号などあればご記入ください（信頼感アップに繋がります）", height=50)

# ===========================================
# 03_選ばれる理由
# ===========================================
ws = wb.create_sheet("03_選ばれる理由")
row = setup_sheet(ws, "選ばれる理由", "会社の強み")

strengths = [
    ("有資格者による確かな品質", "一級土木施工管理技士が現場を管理。計画から施工まで責任を持って対応します。"),
    ("解体から活用までワンストップ", "解体・土地開発・不動産・リフォームを自社で一貫対応。窓口ひとつで完結します。"),
    ("地域密着・スピード対応", "北九州エリアを中心にフットワーク軽く対応。お見積りは無料でご相談いただけます。"),
]
for i, (title, desc) in enumerate(strengths, start=1):
    row = add_section(ws, row, f"強み{i}　{title}")
    row = add_row(ws, row, "説明文", desc)
    row = add_row(ws, row, "修正・補足", "", height=36)
    row = add_spacer(ws, row)

row = add_section(ws, row, "アピールしたいこと（自由記入）")
row = add_row(ws, row, "他社に負けない点", "", height=50)
row = add_row(ws, row, "お客様の声・実績", "", height=50)

# ===========================================
# 04_会社概要
# ===========================================
ws = wb.create_sheet("04_会社概要ほか")
row = setup_sheet(ws, "会社概要ほか", "会社情報・写真・デザイン")

row = add_section(ws, row, "名刺から掲載済みの情報（確認をお願いします）")
row = add_row(ws, row, "会社名", "株式会社ヴィレッジ")
row = add_row(ws, row, "代表者", "代表取締役　中川 基次")
row = add_row(ws, row, "資格", "一級土木施工管理技士")
row = add_row(ws, row, "所在地", "〒800-0042　福岡県北九州市門司区上馬寄1丁目8-1-1203")
row = add_row(ws, row, "電話番号", "090-9566-9563")
row = add_row(ws, row, "メール", "village2024@outlook.jp")
row = add_row(ws, row, "事業内容", "建物解体工事 / 解体業コンサルタント / 土地開発 / 不動産売買 / リフォーム事業 / リノベーション事業")

row = add_spacer(ws, row)
row = add_section(ws, row, "掲載したい追加情報（分かる範囲でご記入ください）")
row = add_row(ws, row, "設立・創業年", "※ 未確認（例：2024年 設立）")
row = add_row(ws, row, "資本金", "※ 未確認（掲載しない選択も可）")
row = add_row(ws, row, "従業員数", "※ 未確認（掲載しない選択も可）")
row = add_row(ws, row, "FAX番号", "※ 未確認")
row = add_row(ws, row, "営業時間", "（ドラフト）8:00〜19:00（日曜・祝日を除く）　※ 実際の時間をご記入ください")
row = add_row(ws, row, "定休日", "（ドラフト）日曜・祝日　※ 変更あればご記入ください")
row = add_row(ws, row, "建設業許可番号", "※ あればご記入ください")
row = add_row(ws, row, "宅地建物取引業免許", "※ 不動産売買の免許番号があればご記入ください")
row = add_row(ws, row, "取引銀行", "", height=30)
row = add_row(ws, row, "沿革・会社の歴史", "", height=60)

# --- ご提供写真（旧06を統合） ---
row = add_spacer(ws, row)
row = add_section(ws, row, "ご提供いただきたい写真（右欄にメモをご記入ください）")
photos = [
    ("トップ背景（一番目立つ写真）", "解体現場・重機・整地後の土地など、迫力のある横長写真がおすすめ"),
    ("代表者の写真", "中川様のお顔写真（スーツ・作業着どちらでも）※任意"),
    ("会社ロゴのデータ", "名刺のVマーク。画像データ（PNG/AI等）があれば高画質で掲載できます"),
    ("解体工事の写真", "施工前・施工中・施工後など（あるだけ）"),
    ("重機・車両の写真", "自社の重機・トラックなど"),
    ("リフォーム・リノベ事例", "ビフォーアフターがあると効果的"),
    ("不動産・土地の写真", "取扱い物件・造成地など"),
    ("スタッフ・作業風景", "働く様子が伝わる写真"),
    ("会社外観・事務所", "外観や看板など"),
]
for name, detail in photos:
    r = row
    lc = ws.cell(row=r, column=2, value=name)
    lc.font = LABEL_FONT; lc.fill = LABEL_FILL; lc.alignment = LEFT_WRAP; lc.border = BORDER
    cc = ws.cell(row=r, column=3, value=detail)
    cc.font = CURRENT_FONT; cc.fill = CURRENT_FILL; cc.alignment = LEFT_WRAP; cc.border = BORDER
    ic = ws.cell(row=r, column=4, value="□ 提供できる　／　□ 後日用意　／　□ 撮影してほしい　／　□ 不要")
    ic.font = INPUT_FONT; ic.fill = INPUT_FILL; ic.alignment = LEFT_WRAP; ic.border = BORDER
    ws.row_dimensions[r].height = 34
    row += 1
c = ws.cell(row=row, column=2, value="※ スマートフォンで撮影した写真でも構いません。データはメール添付やギガファイル便等でお送りください。")
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
c.font = NOTE_FONT; c.alignment = LEFT_WRAP
ws.row_dimensions[row].height = 22
row += 1

# --- デザインのご希望・参考URL（旧07を統合） ---
row = add_spacer(ws, row)
row = add_section(ws, row, "サイトの色・雰囲気")
row = add_row(ws, row, "メインカラー（ドラフト）", "青 × 緑（名刺のVマークの色に合わせています）")
row = add_row(ws, row, "色・雰囲気のご希望",
    "※ 変更したい色や目指したい印象をご記入ください。\n（例：信頼感 / 誠実 / 力強い / 清潔感 / 親しみやすい / シンプル 等）", height=48)

row = add_spacer(ws, row)
row = add_section(ws, row, "参考にしたいサイト（あればURLをご記入ください）")
row = add_row(ws, row, "参考URL①", "URL：\n良いと思った点：", height=44)
row = add_row(ws, row, "参考URL②", "URL：\n良いと思った点：", height=44)
row = add_row(ws, row, "その他ご要望・ご質問", "", height=60)

# ===========================================
# 05_お問い合わせ
# ===========================================
ws = wb.create_sheet("05_お問い合わせ")
row = setup_sheet(ws, "お問い合わせ", "連絡先の表示")

row = add_section(ws, row, "電話")
row = add_row(ws, row, "電話番号", "090-9566-9563")
row = add_row(ws, row, "受付時間の表記", "（ドラフト）8:00〜19:00（日曜・祝日を除く）")

row = add_spacer(ws, row)
row = add_section(ws, row, "メール")
row = add_row(ws, row, "メールアドレス", "village2024@outlook.jp")
row = add_row(ws, row, "備考", "※ お問い合わせ専用アドレスをご希望の場合は作成も可能です")

row = add_spacer(ws, row)
row = add_section(ws, row, "お問い合わせフォームについて")
row = add_row(ws, row, "フォーム設置", "※ サイトから直接送信できる入力フォームを設置しますか？（はい / いいえ）", height=36)
row = add_row(ws, row, "受信先メール", "※ フォームを設置する場合、受信したいメールアドレスをご記入ください")
row = add_row(ws, row, "地図の表示", "※ Googleマップ（所在地）を掲載しますか？（はい / いいえ）", height=36)

row = add_spacer(ws, row)
row = add_section(ws, row, "ホームページのアドレス（ドメイン）")
row = add_row(ws, row, "ご希望のドメイン",
    "※ ホームページのURLに使いたい文字列のご希望があればご記入ください。\n"
    "　（例：village-kitakyushu.jp / village-kaitai.com など）\n"
    "　取得できるか確認し、こちらで取得手続きを行います。特に希望が無ければお任せでも大丈夫です。", height=64)
row = add_row(ws, row, "希望する末尾", "※ ご希望があれば○：　.jp　/　.com　/　.co.jp　/　こだわらない", height=36)

# ===========================================
# 保存
# ===========================================
out = r"c:\Users\user\Desktop\village\docs\ヴィレッジHP_ヒアリングシート.xlsx"
wb.save(out)
print(f"OK: {out}")
